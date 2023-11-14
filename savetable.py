import openpyxl
import xlsxwriter
import pandas as pd
import os
from copy import copy
from openpyxl.utils.dataframe import dataframe_to_rows
def savetable(input,output,df):
    if not os.path.isfile(input):
        wb = openpyxl.Workbook()
        sheet=wb.active
        wb.save(input)
    source_book = openpyxl.load_workbook(input)
    # Загрузка исходного листа
    source_sheet = source_book['Sheet']
    # Создание новой книги
    target_book = openpyxl.Workbook()
    # Создание нового листа
    target_sheet = target_book.active
    ws=target_book.active
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    # Перенос стилей ячеек
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet[cell.coordinate]
            target_cell.font = copy(cell.font)
            target_cell.border = copy(cell.border)
            target_cell.fill = copy(cell.fill)
            target_cell.number_format = copy(cell.number_format)
            target_cell.protection = copy(cell.protection)
            target_cell.alignment = copy(cell.alignment)
    target_book.save(output)
